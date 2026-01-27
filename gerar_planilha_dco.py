#!/usr/bin/env python3
"""
Gera a planilha DCO para a campanha Uniforme Escolar do GDF.
Estrutura: 3 personagens x 2 subtexts_frame2 x 4 formatos = 24 combinacoes
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Dados da campanha
PERSONAGENS = [
    {"id": "menina01", "nome": "Menina com trancas"},
    {"id": "menino01", "nome": "Menino com oculos"},
    {"id": "menino02", "nome": "Menino loiro"},
]

# Headline fixo
HEADLINE = "Cartao Uniforme Escolar."

# Subtext Frame 1 fixo
SUBTEXT_FRAME1 = "Feito na medida certa para 442 mil estudantes das escolas publicas."

# Opcoes de Subtext Frame 2
SUBTEXTS_FRAME2 = [
    "Desbloqueie o seu cartao no aplicativo BRB Social e confira as malharias credenciadas.",
    "Em caso de duvidas, procure a Regional de Ensino do seu filho.",
]

FORMATOS = [
    {"nome": "300x250", "width": 300, "height": 250, "tipo": "Medium Rectangle"},
    {"nome": "468x60", "width": 468, "height": 60, "tipo": "Full Banner"},
    {"nome": "728x90", "width": 728, "height": 90, "tipo": "Leaderboard"},
    {"nome": "970x250", "width": 970, "height": 250, "tipo": "Billboard"},
]

BASE_URL = "https://agenciabrasilia.df.gov.br/w/cartao-uniforme-escolar-permite-a-aquisicao-de-itens-para-estudantes-da-rede-publica-em-malharias-credenciadas?redirect=%2Fnoticias&utm_source=dv360&utm_medium=display&utm_campaign=uniforme_escolar&utm_content={reporting_label}"


def create_workbook():
    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0C326F", end_color="0C326F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aba 1: Feed Principal
    ws_feed = wb.active
    ws_feed.title = "Feed_Principal"

    headers = [
        "ID", "Reporting_Label", "Formato", "Width", "Height",
        "Personagem_ID", "Personagem_Asset",
        "Headline", "Subtext_Frame1", "Subtext_Frame2",
        "ExitURL", "Default", "Active"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws_feed.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Gerar todas as combinacoes
    row_id = 0
    for formato in FORMATOS:
        for p_idx, personagem in enumerate(PERSONAGENS):
            for s_idx, subtext2 in enumerate(SUBTEXTS_FRAME2):
                row_id += 1
                row = row_id + 1

                reporting_label = f"{row_id:03d}_{personagem['id']}_{formato['nome']}_S{s_idx+1}"
                asset = f"{personagem['id']}_{formato['nome']}.png"
                exit_url = BASE_URL.format(reporting_label=reporting_label)
                is_default = "TRUE" if (p_idx == 0 and s_idx == 0) else "FALSE"

                data = [
                    row_id,
                    reporting_label,
                    formato['nome'],
                    formato['width'],
                    formato['height'],
                    personagem['id'],
                    asset,
                    HEADLINE,
                    SUBTEXT_FRAME1,
                    subtext2,
                    exit_url,
                    is_default,
                    "TRUE"
                ]

                for col, value in enumerate(data, 1):
                    cell = ws_feed.cell(row=row, column=col, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

    # Ajustar larguras
    column_widths = [6, 35, 10, 8, 8, 12, 25, 25, 55, 70, 120, 8, 8]
    for col, width in enumerate(column_widths, 1):
        ws_feed.column_dimensions[get_column_letter(col)].width = width

    # Aba 2: Textos
    ws_textos = wb.create_sheet("Textos")

    ws_textos.cell(row=1, column=1, value="Headline (fixo)").font = header_font
    ws_textos.cell(row=1, column=1).fill = header_fill
    ws_textos.merge_cells('A1:B1')
    ws_textos.cell(row=2, column=1, value="H1")
    ws_textos.cell(row=2, column=2, value=HEADLINE)

    ws_textos.cell(row=4, column=1, value="Subtext Frame 1 (fixo)").font = header_font
    ws_textos.cell(row=4, column=1).fill = header_fill
    ws_textos.merge_cells('A4:B4')
    ws_textos.cell(row=5, column=1, value="S1")
    ws_textos.cell(row=5, column=2, value=SUBTEXT_FRAME1)

    ws_textos.cell(row=7, column=1, value="Subtext Frame 2 (variavel)").font = header_font
    ws_textos.cell(row=7, column=1).fill = header_fill
    ws_textos.merge_cells('A7:B7')

    for idx, subtext in enumerate(SUBTEXTS_FRAME2, 1):
        ws_textos.cell(row=7+idx, column=1, value=f"S2_{idx}")
        ws_textos.cell(row=7+idx, column=2, value=subtext)

    ws_textos.column_dimensions['A'].width = 12
    ws_textos.column_dimensions['B'].width = 80

    # Aba 3: Personagens
    ws_personagens = wb.create_sheet("Personagens")

    pers_headers = ["ID", "Nome", "Asset_300x250", "Asset_468x60", "Asset_728x90", "Asset_970x250"]
    for col, header in enumerate(pers_headers, 1):
        cell = ws_personagens.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for idx, pers in enumerate(PERSONAGENS, 2):
        ws_personagens.cell(row=idx, column=1, value=pers['id'])
        ws_personagens.cell(row=idx, column=2, value=pers['nome'])
        ws_personagens.cell(row=idx, column=3, value=f"{pers['id']}_300x250.png")
        ws_personagens.cell(row=idx, column=4, value=f"{pers['id']}_468x60.png")
        ws_personagens.cell(row=idx, column=5, value=f"{pers['id']}_728x90.png")
        ws_personagens.cell(row=idx, column=6, value=f"{pers['id']}_970x250.png")

    for col in range(1, 7):
        ws_personagens.column_dimensions[get_column_letter(col)].width = 22

    # Aba 4: Formatos
    ws_formatos = wb.create_sheet("Formatos")

    format_headers = ["Formato", "Width", "Height", "Tipo"]
    for col, header in enumerate(format_headers, 1):
        cell = ws_formatos.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for idx, fmt in enumerate(FORMATOS, 2):
        ws_formatos.cell(row=idx, column=1, value=fmt['nome'])
        ws_formatos.cell(row=idx, column=2, value=fmt['width'])
        ws_formatos.cell(row=idx, column=3, value=fmt['height'])
        ws_formatos.cell(row=idx, column=4, value=fmt['tipo'])

    for col in range(1, 5):
        ws_formatos.column_dimensions[get_column_letter(col)].width = 18

    return wb


if __name__ == "__main__":
    import os

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "DCO_GDF_UNIFORME_ESCOLAR.xlsx")

    wb = create_workbook()
    wb.save(output_path)

    total = len(PERSONAGENS) * len(SUBTEXTS_FRAME2) * len(FORMATOS)
    print(f"Planilha gerada: {output_path}")
    print(f"Total de combinacoes no Feed Principal: {total} linhas")
